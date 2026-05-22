import openpyxl

def convert_to_tab_delimited_single_column():
    input_file = "Excel-TSV-Input.xlsx"
    output_file = "Excel-TSV-Output.csv"
    
    # Use \t for the Tab character
    delimiter = "\t" 

    try:
        # Load workbook
        wb = openpyxl.load_workbook(input_file, data_only=True)
        sheet = wb.active

        # Determine the absolute boundaries
        max_row = sheet.max_row
        max_col = sheet.max_column

        with open(output_file, 'w', encoding='utf-8', newline='') as f:
            for r in range(1, max_row + 1):
                row_data = []
                for c in range(1, max_col + 1):
                    val = sheet.cell(row=r, column=c).value
                    
                    # Convert to string and clean out any internal line breaks
                    if val is not None:
                        clean_val = str(val).replace('\n', ' ').replace('\r', '')
                    else:
                        clean_val = ""
                        
                    row_data.append(clean_val)
                
                # Join fields with a Tab
                combined_row = delimiter.join(row_data)
                
                # Wrap in double quotes so it lands in Column A
                f.write(f'"{combined_row}"\n')

        print(f"Success! Created '{output_file}' using Tabs as the separator.")
        print(f"Grid size: {max_row} rows by {max_col} columns.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    convert_to_tab_delimited_single_column()
